import configparser

import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import Color, PatternFill
from openpyxl.cell import Cell

from tools.ManualParam import ManualParam
from tools import ManualParam as mp

import os as os
import shutil as shutil
from test.test_funcattrs import cell


class AuditFileController():
    
    outputDir:str = ""
    auditDataFolder = "auditData"
    manualParam:mp = mp.ManualParam()
    wb = openpyxl.Workbook()
    orgFile:str = ""
    auditFileName:str = ""
    currentRows:int = 0
    dataStartRow:int = 9
    
    # コンストラクタ
    def __init__(self):
        print("init:AuditFileController")
        # ログインユーザアカウント取得して、デスクトップに配置する。
        whoAmI:str = os.environ['USERNAME']
        suntoryEnv = configparser.ConfigParser()
        suntoryEnv.read('suntoryEnv.ini', encoding='utf-8')
        #outputBase:str = suntoryEnv['OUTPUT_PATH']['outputPathBase']
        #lastDir:str = suntoryEnv['OUTPUT_PATH']['outputPathDesktop']
        #self.outputDir = outputBase + whoAmI + lastDir
       # self.orgFile = suntoryEnv['ORIGINAL_FILE']['orgFile']
        
        self.currentRows = 9
        
        #self.outputdir = "C:\\ishikawa\work\\suntory_Dev\\";
        #self.wb = pyxl.load_workbook(self.outputdir + "auditDoc.xlsx") 
        #sheet = self.wb.active
        #sheet.title = 'AuditTask'
        
    
    # オリジナルの監査資料ファイルを、実行しているユーザのデスクトップへコピーする。
    def copyOriginalFile(self,fileName):
        print("ベースの監査資料ファイルコピー処理　START")
        self.auditFileName = fileName
        shutil.copyfile(self.orgFile,self.outputDir + self.auditFileName)

    def openTargetAuditDoc(self,mp):
        print("監査資料ｘｌｓｘファイルオープン処理　START")
        manualParam:ManualParam = mp
        
        wb = openpyxl.load_workbook(self.outputDir + self.auditFileName)
        sheet = wb["auditDoc"]
        
        # Selected defect mode(K2セル)書き込み
        sheet.cell(row=2,column=11).value = manualParam.get_selectedDefectMode()
        # Selected packaging(K3セル)書き込み
        sheet.cell(row=3,column=11).value = manualParam.get_selectedPackaging()
        # Selected line(K4セル)書き込み
        sheet.cell(row=4,column=11).value = manualParam.get_selectedLine()
        # Selected Section(K5セル)書き込み
        sheet.cell(row=5,column=11).value = manualParam.get_selectedSection()
        # priorityOnly(K6セル)書き込み
        sheet.cell(row=6,column=11).value = manualParam.get_priorityOnly()
        
        # Subjected defect mode 書き込み(B列)
        sheet.cell(row=self.currentRows,column=2).value = manualParam.get_subjectedDefectMode()
        # Subjected packaging 書き込み(C列)
        sheet.cell(row=self.currentRows,column=3).value = manualParam.get_subjectedPackaging()
        # Subjected line 書き込み(D列)
        sheet.cell(row=self.currentRows,column=4).value = manualParam.get_subjectedLine()
        # Section 書き込み(E列)
        sheet.cell(row=self.currentRows,column=5).value = manualParam.get_section()
        # Subsection 書き込み(F列)
        sheet.cell(row=self.currentRows,column=6).value = manualParam.get_subSection()
        # Subsubsection 書き込み(G列)
        sheet.cell(row=self.currentRows,column=7).value = manualParam.get_subsubSection()
        # Purpose書き込み(H列)
        sheet.cell(row=self.currentRows,column=8).value = manualParam.get_purpose()

        # What/How to achieve書き込み(I列)
        # What/How to achievePriority書き込み(L列)
        # 循環参照用What/How to achieve書き込み(AC列)
        # 循環参照用What/How to achievePriority書き込み(AD列)
        htaCnt:int = 0
        howToAchivePriorityList:list[str] = manualParam.get_howToAchievePriority()
        if len(manualParam.get_howToAchieve()) > 0 :
            for how2AchieveVal in manualParam.get_howToAchieve() :
                # 値書き込み
                sheet.cell(row=self.currentRows + htaCnt,column=9).value = how2AchieveVal
                sheet.cell(row=self.currentRows + htaCnt,column=12).value = howToAchivePriorityList[htaCnt]
                # 循環参照用の値書き込み(AC列,AD列)
                sheet.cell(row=self.currentRows + htaCnt,column=29).value = how2AchieveVal
                sheet.cell(row=self.currentRows + htaCnt,column=30).value = howToAchivePriorityList[htaCnt]
                
                rowHeight:int = self.calcRowHeight(how2AchieveVal)
                sheet.row_dimensions[self.currentRows + htaCnt].height = rowHeight
                
                currentCell:str = 'I' + str(self.currentRows + htaCnt)
                alignment = Alignment(wrap_text=True)
                sheet[currentCell].alignment = alignment
                # 行(=横方面)の結合
                sheet.merge_cells(start_row=self.currentRows + htaCnt, end_row=self.currentRows + htaCnt , start_column=9, end_column=11)

                # 罫線の処理
                side = Side(style='thin', color='000000')
                border = Border(top=side, bottom=side,
                                left=side, right=side)
                sheet.cell(row=self.currentRows + htaCnt,column=9).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=10).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=11).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=12).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=13).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=14).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=15).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=16).border = border
                sheet.cell(row=self.currentRows + htaCnt,column=17).border = border
                
                htaCnt = htaCnt + 1

        self.currentRows = self.currentRows + htaCnt
        
        ### Task1書き込み(K列) ###
        # priorityの条件有無によっては、Taskが一つもない監査資料を生成する可能性もあるので、
        # 各情報の単位で分岐を付与する。（けっこう苦肉）
        # Task1(Procedure)書き込み(K列)
        task1ProcCnt:int = 0
        task1DefaultCell:int = self.currentRows
        
        # Task1の出力件数が1件以上(=0件でない場合)、Task1の文言と各種値、罫線背景色の処理を行う。
        if manualParam.get_task1OutputCnt() > 0 :
            #　Task1固定文言出力
            sheet.cell(row=task1DefaultCell,column=9).value = "Task 1"
            
            # 背景色を変更
            fill = PatternFill(patternType='solid', fgColor='DDEBF7', bgColor='DDEBF7')
            sheet.cell(row=task1DefaultCell,column=9).fill = fill
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task1Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task1Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
                # 罫線の処理
                side = Side(style='thin', color='000000')
                border = Border(top=side, bottom=side, left=side, right=side)
                sheet.cell(row=self.currentRows,column=10).border = border
            
            if len(manualParam.get_task1Procedure()) > 0 :
                for task1ProcVal in manualParam.get_task1Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task1ProcCnt,column=11).value = task1ProcVal
                    if manualParam.get_task1AllPriority() == True :
                        sheet.cell(row=self.currentRows + task1ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task1ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task1ProcCnt = task1ProcCnt + 1
            self.currentRows = self.currentRows + task1ProcCnt
        
            # Task1(Frequency or timing)書き込み(K列)
            task1FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task1Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task1Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
                # 罫線の処理
                side = Side(style='thin', color='000000')
                border = Border(top=side, bottom=side, left=side, right=side)
                sheet.cell(row=self.currentRows,column=10).border = border
                
            if len(manualParam.get_task1Frequency()) > 0 :
                for task1FreqVal in manualParam.get_task1Frequency() :
                    sheet.cell(row=self.currentRows + task1FreqCnt,column=11).value = task1FreqVal
                    if manualParam.get_task1AllPriority() == True :
                        sheet.cell(row=self.currentRows + task1FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task1FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    task1FreqCnt = task1FreqCnt + 1
            self.currentRows = self.currentRows + task1FreqCnt
            
            # Task1(Operating limit or standard condition)書き込み(K列)
            task1OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task1Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task1Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
                # 罫線の処理
                side = Side(style='thin', color='000000')
                border = Border(top=side, bottom=side, left=side, right=side)
                sheet.cell(row=self.currentRows,column=10).border = border
            
            if len(manualParam.get_task1Operating()) > 0 :
                for task1OperateVal in manualParam.get_task1Operating() :
                    sheet.cell(row=self.currentRows + task1OperateCnt,column=11).value = task1OperateVal
                    if manualParam.get_task1AllPriority() == True :
                        sheet.cell(row=self.currentRows + task1OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task1OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task1OperateCnt = task1OperateCnt + 1
                    
            if manualParam.get_task1AllPriority() == False :
                cntTask1Child:int = 0
                for val in manualParam.get_task1ChildPriority() : 
                    sheet.cell(row=task1DefaultCell + cntTask1Child,column=12).value = val
                    cntTask1Child = cntTask1Child + 1 
                    
            self.currentRows = self.currentRows + task1OperateCnt
            
            # Task1の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
            
            tmpTask1Cnt:int = task1DefaultCell
            while tmpTask1Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask1Cnt,column=9).border = border
                tmpTask1Cnt = tmpTask1Cnt + 1
            
            # Task1ヘッダのセル結合
            # task1DefaultCell から currentRows-1(task1Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task1DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)
        
        ### Task2書き込み(K列) ###
        # Task2以降は可変データなので、procedureが存在する場合は当該Taskが存在するとみなして
        # データ出力を行う。
        # Task2(Procedure)書き込み(K列)
        if len(manualParam.get_task2Procedure()) > 0 :
            task2ProcCnt:int = 0
            task2DefaultCell:int = self.currentRows
            sheet.cell(row=task2DefaultCell,column=9).value = "Task 2"
            # 背景色を変更
            sheet.cell(row=task2DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task2Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task2Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task2Procedure()) > 0 :
                for task2ProcVal in manualParam.get_task2Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task2ProcCnt,column=11).value = task2ProcVal
                    if manualParam.get_task2AllPriority() == True :
                        sheet.cell(row=self.currentRows + task2ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task2ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1

                    task2ProcCnt = task2ProcCnt + 1
            self.currentRows = self.currentRows + task2ProcCnt
            
            # Task2(Frequency or timing)書き込み(K列)
            task2FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task2Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task2Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task2Frequency()) > 0 :
                for task2FreqVal in manualParam.get_task2Frequency() :
                    sheet.cell(row=self.currentRows + task2FreqCnt,column=11).value = task2FreqVal
                    if manualParam.get_task2AllPriority() == True :
                        sheet.cell(row=self.currentRows + task2FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task2FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task2FreqCnt = task2FreqCnt + 1
            self.currentRows = self.currentRows + task2FreqCnt
            
            # Task2(Operating limit or standard condition)書き込み(K列)
            task2OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task2Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task2Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task2Operating()) > 0 :
                for task2OperateVal in manualParam.get_task2Operating() :
                    sheet.cell(row=self.currentRows + task2OperateCnt,column=11).value = task2OperateVal
                    if manualParam.get_task2AllPriority() == True :
                        sheet.cell(row=self.currentRows + task2OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task2OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task2OperateCnt = task2OperateCnt + 1
            self.currentRows = self.currentRows + task2OperateCnt
            
            # Task2の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask2Cnt:int = task2DefaultCell
            while tmpTask2Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask2Cnt,column=9).border = border
                tmpTask2Cnt = tmpTask2Cnt + 1
            
            # Task2ヘッダのセル結合
            # task2DefaultCell から currentRows-1(task2Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task2DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)
            
        ### Task3書き込み(K列) ###
        # Task3(Procedure)書き込み(K列)
        if len(manualParam.get_task3Procedure()) > 0 :
            task3ProcCnt:int = 0
            task3DefaultCell:int = self.currentRows
            sheet.cell(row=task3DefaultCell,column=9).value = "Task 3"
            # 背景色を変更
            sheet.cell(row=task3DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task3Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task3Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task3Procedure()) > 0 :
                for task3ProcVal in manualParam.get_task3Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task3ProcCnt,column=11).value = task3ProcVal
                    if manualParam.get_task3AllPriority() == True :
                        sheet.cell(row=self.currentRows + task3ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)

                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task3ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task3ProcCnt = task3ProcCnt + 1
            self.currentRows = self.currentRows + task3ProcCnt
            
            # task3(Frequency or timing)書き込み(K列)
            task3FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task3Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task3Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task3Frequency()) > 0 :
                for task3FreqVal in manualParam.get_task3Frequency() :
                    sheet.cell(row=self.currentRows + task3FreqCnt,column=11).value = task3FreqVal
                    if manualParam.get_task3AllPriority() == True :
                        sheet.cell(row=self.currentRows + task3FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task3FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task3FreqCnt = task3FreqCnt + 1
            self.currentRows = self.currentRows + task3FreqCnt
            
            # task3(Operating limit or standard condition)書き込み(K列)
            task3OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task3Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task3Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task3Operating()) > 0 :
                for task3OperateVal in manualParam.get_task3Operating() :
                    sheet.cell(row=self.currentRows + task3OperateCnt,column=11).value = task3OperateVal
                    if manualParam.get_task3AllPriority() == True :
                        sheet.cell(row=self.currentRows + task3OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task3OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task3OperateCnt = task3OperateCnt + 1
            self.currentRows = self.currentRows + task3OperateCnt
            
            # Task3の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask3Cnt:int = task3DefaultCell
            while tmpTask3Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask3Cnt,column=9).border = border
                tmpTask3Cnt = tmpTask3Cnt + 1
            
            # task3ヘッダのセル結合
            # task3DefaultCell から currentRows-1(task3Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task3DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)

        ### Task4書き込み(K列) ###
        # Task4(Procedure)書き込み(K列)
        if len(manualParam.get_task4Procedure()) > 0 :
            task4ProcCnt:int = 0
            task4DefaultCell:int = self.currentRows
            sheet.cell(row=task4DefaultCell,column=9).value = "Task 4"
            # 背景色を変更
            sheet.cell(row=task4DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task4Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task4Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task4Procedure()) > 0 :
                for task4ProcVal in manualParam.get_task4Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task4ProcCnt,column=11).value = task4ProcVal
                    if manualParam.get_task4AllPriority() == True :
                        sheet.cell(row=self.currentRows + task4ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task4ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task4ProcCnt = task4ProcCnt + 1
            self.currentRows = self.currentRows + task4ProcCnt
            
            # task4(Frequency or timing)書き込み(K列)
            task4FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task4Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task4Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task4Frequency()) > 0 :
                for task4FreqVal in manualParam.get_task4Frequency() :
                    sheet.cell(row=self.currentRows + task4FreqCnt,column=11).value = task4FreqVal
                    if manualParam.get_task4AllPriority() == True :
                        sheet.cell(row=self.currentRows + task4FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task4FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task4FreqCnt = task4FreqCnt + 1
            self.currentRows = self.currentRows + task4FreqCnt
            
            # task4(Operating limit or standard condition)書き込み(K列)
            task4OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task4Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task4Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task4Operating()) > 0 :
                for task4OperateVal in manualParam.get_task4Operating() :
                    sheet.cell(row=self.currentRows + task4OperateCnt,column=11).value = task4OperateVal
                    if manualParam.get_task4AllPriority() == True :
                        sheet.cell(row=self.currentRows + task4OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task4OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task4OperateCnt = task4OperateCnt + 1
            self.currentRows = self.currentRows + task4OperateCnt
            
            # Task4の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask4Cnt:int = task4DefaultCell
            while tmpTask4Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask4Cnt,column=9).border = border
                tmpTask4Cnt = tmpTask4Cnt + 1
            
            # task4ヘッダのセル結合
            # task4DefaultCell から currentRows-1(task4Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task4DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)

        ### Task5書き込み(K列) ###
        # Task5(Procedure)書き込み(K列)
        if len(manualParam.get_task5Procedure()) > 0 :
            task5ProcCnt:int = 0
            task5DefaultCell:int = self.currentRows
            sheet.cell(row=task5DefaultCell,column=9).value = "Task 5"
            # 背景色を変更
            sheet.cell(row=task5DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task5Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task5Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task5Procedure()) > 0 :
                for task5ProcVal in manualParam.get_task5Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task5ProcCnt,column=11).value = task5ProcVal
                    if manualParam.get_task5AllPriority() == True :
                        sheet.cell(row=self.currentRows + task5ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task5ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task5ProcCnt = task5ProcCnt + 1
            self.currentRows = self.currentRows + task5ProcCnt
            
            # task5(Frequency or timing)書き込み(K列)
            task5FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task5Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task5Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task5Frequency()) > 0 :
                for task5FreqVal in manualParam.get_task5Frequency() :
                    sheet.cell(row=self.currentRows + task5FreqCnt,column=11).value = task5FreqVal
                    if manualParam.get_task5AllPriority() == True :
                        sheet.cell(row=self.currentRows + task5FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task5FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task5FreqCnt = task5FreqCnt + 1
            self.currentRows = self.currentRows + task5FreqCnt
            
            # task5(Operating limit or standard condition)書き込み(K列)
            task5OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task5Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task5Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task5Operating()) > 0 :
                for task5OperateVal in manualParam.get_task5Operating() :
                    sheet.cell(row=self.currentRows + task5OperateCnt,column=11).value = task5OperateVal
                    if manualParam.get_task5AllPriority() == True :
                        sheet.cell(row=self.currentRows + task5OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task5OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task5OperateCnt = task5OperateCnt + 1
            self.currentRows = self.currentRows + task5OperateCnt
            
            # Task5の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask5Cnt:int = task5DefaultCell
            while tmpTask5Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask5Cnt,column=9).border = border
                tmpTask5Cnt = tmpTask5Cnt + 1
            
            # task5ヘッダのセル結合
            # task5DefaultCell から currentRows-1(task5Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task5DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)
        
        ### Task6書き込み(K列) ###
        # Task6(Procedure)書き込み(K列)
        if len(manualParam.get_task6Procedure()) > 0 :
            task6ProcCnt:int = 0
            task6DefaultCell:int = self.currentRows
            sheet.cell(row=task6DefaultCell,column=9).value = "Task 6"
            # 背景色を変更
            sheet.cell(row=task6DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task6Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task6Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task6Procedure()) > 0 :
                for task6ProcVal in manualParam.get_task6Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task6ProcCnt,column=11).value = task6ProcVal
                    if manualParam.get_task6AllPriority() == True :
                        sheet.cell(row=self.currentRows + task6ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task6ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task6ProcCnt = task6ProcCnt + 1
            self.currentRows = self.currentRows + task6ProcCnt
            
            # task6(Frequency or timing)書き込み(K列)
            task6FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task6Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task6Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task6Frequency()) > 0 :
                for task6FreqVal in manualParam.get_task6Frequency() :
                    sheet.cell(row=self.currentRows + task6FreqCnt,column=11).value = task6FreqVal
                    if manualParam.get_task6AllPriority() == True :
                        sheet.cell(row=self.currentRows + task6FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task6FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task6FreqCnt = task6FreqCnt + 1
            self.currentRows = self.currentRows + task6FreqCnt
            
            # task6(Operating limit or standard condition)書き込み(K列)
            task6OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task6Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task6Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task6Operating()) > 0 :
                for task6OperateVal in manualParam.get_task6Operating() :
                    sheet.cell(row=self.currentRows + task6OperateCnt,column=11).value = task6OperateVal
                    if manualParam.get_task6AllPriority() == True :
                        sheet.cell(row=self.currentRows + task6OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task6OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task6OperateCnt = task6OperateCnt + 1
                        
            self.currentRows = self.currentRows + task6OperateCnt
            
            # Task6の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask6Cnt:int = task6DefaultCell
            while tmpTask6Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask6Cnt,column=9).border = border
                tmpTask6Cnt = tmpTask6Cnt + 1
            
            # task6ヘッダのセル結合
            # task6DefaultCell から currentRows-1(task6Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task6DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)
        
        ### Task7書き込み(K列) ###
        # Task7(Procedure)書き込み(K列)
        if len(manualParam.get_task7Procedure()) > 0 :
            task7ProcCnt:int = 0
            task7DefaultCell:int = self.currentRows
            sheet.cell(row=task7DefaultCell,column=9).value = "Task 7"
            # 背景色を変更
            sheet.cell(row=task7DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task7Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task7Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task7Procedure()) > 0 :
                for task7ProcVal in manualParam.get_task7Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task7ProcCnt,column=11).value = task7ProcVal
                    if manualParam.get_task7AllPriority() == True :
                        sheet.cell(row=self.currentRows + task7ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task7ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task7ProcCnt = task7ProcCnt + 1
            self.currentRows = self.currentRows + task7ProcCnt
            
            # task7(Frequency or timing)書き込み(K列)
            task7FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task7Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task7Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task7Frequency()) > 0 :
                for task7FreqVal in manualParam.get_task7Frequency() :
                    sheet.cell(row=self.currentRows + task7FreqCnt,column=11).value = task7FreqVal
                    if manualParam.get_task7AllPriority() == True :
                        sheet.cell(row=self.currentRows + task7FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task7FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task7FreqCnt = task7FreqCnt + 1
            self.currentRows = self.currentRows + task7FreqCnt
            
            # task7(Operating limit or standard condition)書き込み(K列)
            task7OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task7Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task7Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task7Operating()) > 0 :
                for task7OperateVal in manualParam.get_task7Operating() :
                    sheet.cell(row=self.currentRows + task7OperateCnt,column=11).value = task7OperateVal
                    if manualParam.get_task7AllPriority() == True :
                        sheet.cell(row=self.currentRows + task7OperateCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task7OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task7OperateCnt = task7OperateCnt + 1
            self.currentRows = self.currentRows + task7OperateCnt
            
            # Task7の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask7Cnt:int = task7DefaultCell
            while tmpTask7Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask7Cnt,column=9).border = border
                tmpTask7Cnt = tmpTask7Cnt + 1

            # task7ヘッダのセル結合
            # task7DefaultCell から currentRows-1(task7Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task7DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)
        
        ### Task8書き込み(K列) ###
        # Task8(Procedure)書き込み(K列)
        if len(manualParam.get_task8Procedure()) > 0 :
            task8ProcCnt:int = 0
            task8DefaultCell:int = self.currentRows
            sheet.cell(row=task8DefaultCell,column=9).value = "Task 8"
            # 背景色を変更
            sheet.cell(row=task8DefaultCell,column=9).fill = fill
            
            sheet.cell(row=self.currentRows,column=10).value = "Procedure"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task8Procedure()) > 1 :
                # Procedureのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task8Procedure()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task8Procedure()) > 0 :
                for task8ProcVal in manualParam.get_task8Procedure() :
                    # procedureのデータ書き込み
                    sheet.cell(row=self.currentRows + task8ProcCnt,column=11).value = task8ProcVal
                    if manualParam.get_task8AllPriority() == True :
                        sheet.cell(row=self.currentRows + task8ProcCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task8ProcCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task8ProcCnt = task8ProcCnt + 1
            self.currentRows = self.currentRows + task8ProcCnt
            
            # task8(Frequency or timing)書き込み(K列)
            task8FreqCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Frequency or timing"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            
            if len(manualParam.get_task8Frequency()) > 1 :
                # Frequency or timingのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task8Frequency()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task8Frequency()) > 0 :
                for task8FreqVal in manualParam.get_task8Frequency() :
                    sheet.cell(row=self.currentRows + task8FreqCnt,column=11).value = task8FreqVal
                    if manualParam.get_task8AllPriority() == True :
                        sheet.cell(row=self.currentRows + task8FreqCnt,column=12).value = "ON"
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task8FreqCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task8FreqCnt = task8FreqCnt + 1
            self.currentRows = self.currentRows + task8FreqCnt
            
            # task8(Operating limit or standard condition)書き込み(K列)
            task8OperateCnt:int = 0
            sheet.cell(row=self.currentRows,column=10).value = "Operating limit or standard condition"
            # 背景色変更
            sheet.cell(row=self.currentRows,column=10).fill = fill
            if len(manualParam.get_task8Operating()) > 1 :
                # Operating limit or standard conditionのセル結合
                mcEndInt:int = self.currentRows + len(manualParam.get_task8Operating()) -1
                sheet.merge_cells(start_row=self.currentRows, end_row=mcEndInt , start_column=10, end_column=10)
            
            if len(manualParam.get_task8Operating()) > 0 :
                for task8OperateVal in manualParam.get_task8Operating() :
                    sheet.cell(row=self.currentRows + task8OperateCnt,column=11).value = task8OperateVal
                    side = Side(style='thin', color='000000')
                    border = Border(top=side, bottom=side, left=side, right=side)
                    
                    colTmp:int = 10
                    while colTmp <= 17 :
                        sheet.cell(row=self.currentRows + task8OperateCnt,column=colTmp).border = border
                        colTmp = colTmp + 1
                    
                    task8OperateCnt = task8OperateCnt + 1
            self.currentRows = self.currentRows + task8OperateCnt
            
            # Task8の罫線描画
            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)
        
            tmpTask8Cnt:int = task8DefaultCell
            while tmpTask8Cnt <= self.currentRows-1 :
                sheet.cell(row=tmpTask8Cnt,column=9).border = border
                tmpTask8Cnt = tmpTask8Cnt + 1

            # task8ヘッダのセル結合
            # task8DefaultCell から currentRows-1(task8Operateの書き込みが終わったカウンタ)までを結合する。
            sheet.merge_cells(start_row=task8DefaultCell, end_row=self.currentRows-1 , start_column=9, end_column=9)
        
        # 循環参照用数式の書き込み(V列)
        circulationFormular4V:str = "=$B$" + str(self.dataStartRow)
        cnt4CirculationFormular:int = 0
        cnt4CirculationFormular = self.dataStartRow
        # V列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=22).value = circulationFormular4V
            cnt4CirculationFormular = cnt4CirculationFormular + 1
        
        # ループ終了に伴う初期化
        cnt4CirculationFormular = 0
        # 循環参照用数式の書き込み(W列)
        circulationFormular4W:str = "=$C$" + str(self.dataStartRow)
        cnt4CirculationFormular = self.dataStartRow
        # W列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=23).value = circulationFormular4W
            cnt4CirculationFormular = cnt4CirculationFormular + 1
        
        # ループ終了に伴う初期化
        cnt4CirculationFormular = 0
        
        # 循環参照用数式の書き込み(X列)
        circulationFormular4X:str = "=$D$" + str(self.dataStartRow)
        cnt4CirculationFormular = self.dataStartRow
        # X列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=24).value = circulationFormular4X
            cnt4CirculationFormular = cnt4CirculationFormular + 1
            
        # ループ終了に伴う初期化
        cnt4CirculationFormular = 0
        
        # 循環参照用数式の書き込み(Y列)
        circulationFormular4Y:str = "=$E$" + str(self.dataStartRow)
        cnt4CirculationFormular = self.dataStartRow
        # Y列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=25).value = circulationFormular4Y
            cnt4CirculationFormular = cnt4CirculationFormular + 1
            
        # ループ終了に伴う初期化
        cnt4CirculationFormular = 0
        
        # 循環参照用数式の書き込み(Z列)
        circulationFormular4Z:str = "=$F$" + str(self.dataStartRow)
        cnt4CirculationFormular = self.dataStartRow
        # Z列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=26).value = circulationFormular4Z
            cnt4CirculationFormular = cnt4CirculationFormular + 1
        
        # ループ終了に伴う初期化
        cnt4CirculationFormular = 0
        
        # 循環参照用数式の書き込み(AA列)
        circulationFormular4AA:str = "=$G$" + str(self.dataStartRow)
        cnt4CirculationFormular = self.dataStartRow
        # AA列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=27).value = circulationFormular4AA
            cnt4CirculationFormular = cnt4CirculationFormular + 1
            
        # ループ終了に伴う初期化
        cnt4CirculationFormular = 0
        
        # 循環参照用数式の書き込み(AB列)
        circulationFormular4AB:str = "=$H$" + str(self.dataStartRow)
        cnt4CirculationFormular = self.dataStartRow
        # AA列のループ
        while cnt4CirculationFormular <= self.currentRows-1 :
            sheet.cell(row=cnt4CirculationFormular,column=28).value = circulationFormular4AB
            cnt4CirculationFormular = cnt4CirculationFormular + 1
        
        cnt4CirculationFormular = self.dataStartRow
        
        # 先頭データのセル結合
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=2, end_column=2)
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=3, end_column=3)
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=4, end_column=4)
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=5, end_column=5)
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=6, end_column=6)
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=7, end_column=7)
        sheet.merge_cells(start_row=self.dataStartRow, end_row=self.currentRows-1 , start_column=8, end_column=8)
        
        # B列～H列の罫線描画処理
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        
        colTmpRow:int = self.dataStartRow
        while colTmpRow <= self.currentRows-1 :
            sheet.cell(row=colTmpRow,column=2).border = border
            sheet.cell(row=colTmpRow,column=3).border = border
            sheet.cell(row=colTmpRow,column=4).border = border
            sheet.cell(row=colTmpRow,column=5).border = border
            sheet.cell(row=colTmpRow,column=6).border = border
            sheet.cell(row=colTmpRow,column=7).border = border
            sheet.cell(row=colTmpRow,column=8).border = border
            colTmpRow = colTmpRow + 1
        
        self.dataStartRow = self.currentRows
        
        # 保存
        wb.save(self.outputDir + self.auditFileName)
        
        
    # 文字列の内容によって行幅を変える
    # 「・」の数だけでここでは判定する。
    def calcRowHeight(self,hta):
        staDot:str = "・"
        dotCount:int = 0
        dotCount = hta.count(staDot,0,len(hta))
        if dotCount == 0 :
            if len(hta) > 0 and len(hta) <= 300 :
                return 75
            elif len(hta) > 300 and len(hta) <= 600 :
                return 90
            elif len(hta) > 600 :
                return 105
        elif dotCount == 1 :
            if len(hta) > 0 and len(hta) <= 300 :
                return 75
            elif len(hta) > 300 and len(hta) <= 600 :
                return 90
            elif len(hta) > 600 :
                return 105
        elif dotCount == 2 :
            if len(hta) > 0 and len(hta) <= 300 :
                return 95
            elif len(hta) > 300 and len(hta) <= 600 :
                return 110
            elif len(hta) > 600 :
                return 125
        elif dotCount == 3 :
            if len(hta) > 0 and len(hta) <= 300 :
                return 115
            elif len(hta) > 300 and len(hta) <= 600 :
                return 130
            elif len(hta) > 600 :
                return 145
        elif dotCount == 4 :
            if len(hta) > 0 and len(hta) <= 300 :
                return 135
            elif len(hta) > 300 and len(hta) <= 600 :
                return 150
            elif len(hta) > 600 :
                return 165
